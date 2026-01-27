import requests
from requests.auth import HTTPBasicAuth
from datetime import datetime, timedelta
import json
import csv
from openpyxl import Workbook
from openpyxl.styles import PatternFill

import smtplib
from email.message import EmailMessage
from pathlib import Path

print(datetime.now().strftime('%Y-%m-%d %H:%M:%S'),"-------------Started-------------")
# Email Configuration
SMTP_SERVER = "10.56.131.8"
SMTP_PORT = 25
EMAIL_SENDER = "devopsvmax@airtel.com"
EMAIL_RECIPIENTS = ["Parshant.Verma@airtel.com", "a_Soumyaranjan.Behera@airtel.com", "rajdeepak.bisht@airtel.com", "pawan4.kumar@airtel.com", "Mohammed.A@airtel.com", "Varsha1.Yadav@airtel.com", "Mohd.Ahtasham@airtel.com", "arpit.arora@airtel.com" ,"sanket.arora@airtel.com" ,"Pawan22.Kumar@airtel.com" ,"Chelsi.Bhardwaj@airtel.com" ,"Deeksha.Goel@airtel.com"]
#EMAIL_RECIPIENTS = ["rajdeepak.bisht@airtel.com"]

# Druid endpoint
#url = "http://localhost:8888/druid/v2/sql"
url = "http://10.165.24.8:8888/druid/v2/sql"

# Authentication
username = "admin"
password = "A1GRSL1Uadmin"

"""inventory_metadata = {
"Top Buy Rail":"'62f12e29','16e90cbd','b944b0c0','05e946f8','c7a6d5b0','a745b6a5', \
                '3904dc58','f1ef8e51','7b47ccdc','bbb71fcd','1da29207','bb92672d', \
                'dde95f5a','0fca128e','5a5f3453','eb8e4e16','7735fbf3','61a49bda', \
                'e8ed88e9','3c2faaaa','f2ab0e1f','f730a97e','f76832de','9faa57d8'", \
"OTT Strip":"'b51998a1','5a859e81','160a9f3b','42561e21','411f58a2','4a6b8a53','e93e32e8', \
            '269d1afb','5d66bd98','5e8f9f0f','156dadd9','9bc959bc','87bb88f4','b24dc8f0'", \
"Manage Money Card":"'533ccf4f','91d39b12'", \
"Half Money Card":"'b48bfdab','f6ebe5fd'", \
"Manage proposition":"'c4ee16c7','d32eafb4'", \
"Manage offer card":"'acf85d85','dbff6d13','4b407082','3c474014','5c80c9f1','2b87f967'", \
"Manage Growth Banner":"'5a2b7efa','a972a495','428a4844','1c300cd6','a647f9fd','cfaa505b', \
                        '23563bc0','d4436bad','dc760ab7','21def1a4','4f8c8dea','487ae007'", \
"prepaid details offer card":"'c6a8a89e','b1af9808','a32fa3ce','b9062d0d'", \
"Scratch Card":"'00c93433','a71645b0','f2970a4d','79ccd943','26570824','c7aac47b','ff919627','202ae67c'", \
"XStream":"'37f5eef0','d0aa7bda','3fa197e5','57d5df38','a0c08f55','a7ad4b4c','39c9deef','4eceee79', \
            '49a32a60','8086928','c0621ee1','55ec3176','da81603d','50b81b21','40f2de66','f38c15ca', \
            '0ed6d344','20198259','77ba247f','8c90c103'", \
"Bottom Buy Rail":"'a5b059f8','f7b4a020','ae8885c8','a8d648b1','d858d426','0f9056b9','b94b478f', \
                   'bf2a64e3','378f7ec0','7cfaa7f4','d37fc73e','0d14844d','d81ed45d','c3ed7829', \
                   '2449ee47','aa5fd8ff','98946904','4ceea56b','c77db573','e1851fdc','f4df6d02', \
                   '36db7e46','648e870d','de5fe26f','53f68d4d','8ad045ec','6b244373','01d8b781', \
                   '3eb1f311','86010732'"


                }"""

inventory_metadata = {
    "OTT STRIP":"'87bb88f4','e93e32e8','5d66bd98','156dadd9','160a9f3b','b51998a1','411f58a2','b24dc8f0','9bc959bc','5e8f9f0f','269d1afb','4a6b8a53','42561e21','5a859e81'",
    "MONEY HALF CARD":"'f6ebe5fd','b48bfdab'",
    "MANAGE GROWTH BANNER":"'23563bc0','d4436bad','4f8c8dea','dc760ab7','21def1a4','487ae007','5a2b7efa','a972a495','a647f9fd','428a4844','1c300cd6','cfaa505b'",
    "Prepaid Service Landing Banner":"'ebc988a4','ec0278ea','bc372f2b','bd64f8d1','991c1d0e','3437fb8b','47610088','94befbd4','ebcda6d8','aecafefb','81dfbde1','230b326d'",
    "Scratch Card":"'00c93433','a71645b0'",
    "Data Loan Scratch Card":"'f2970a4d','79ccd943'",
    "Manage Scratch Card":"'c7aac47b','26570824'",
    "FS Top Rail":"'62f12e29','16e90cbd','b944b0c0','05e946f8','c7a6d5b0','a745b6a5','3904dc58','f1ef8e51'",
    "Top Rail":"'bb92672d','1da29207','bbb71fcd','7b47ccdc','eb8e4e16','5a5f3453','0fca128e','dde95f5a'",
    "Financial Services Bottom Rail":"'7735fbf3','61a49bda','e8ed88e9','3c2faaaa','f2ab0e1f','f730a97e','f76832de','9faa57d8'",
    "App Bottom Rail":"'a5b059f8','aa5fd8ff','f7b4a020','98946904','ae8885c8','4ceea56b','a8d648b1','c77db573','d858d426','e1851fdc','0f9056b9','f4df6d02','b94b478f','36db7e46', \
        'bf2a64e3','648e870d','378f7ec0','de5fe26f','7cfaa7f4','53f68d4d','d37fc73e','8ad045ec','0d14844d','6b244373','d81ed45d','01d8b781','c3ed7829','3eb1f311','2449ee47','86010732'",
    "Recharge Shortcut Banner":"'f88469d9','0558a1da','90dc79c9','22ff2261','16a879e0','40b99908'",
    "Manage Proposition":"'c4ee16c7','d32eafb4','22ebce3e','a37d378c'",
    "Manage Money Card":"'533ccf4f','91d39b12'",
    "Service Landing Banner":"'e3b9cb42','6126c258','5eb792e8','cc8b9683','88d7b6b1','2fb90633'",
    "Service Landing Offer":"'21f73db4','335f73ee','07be42bf','19dbbe14','c1d940ae','2bcf730f'",
    "Prepaid Service Landing Banner":"'230b326d','81dfbde1','aecafefb','ebcda6d8','94befbd4','47610088','3437fb8b','991c1d0e','bd64f8d1','bc372f2b','ec0278ea','ebc988a4'",
    "Postpaid Top Banner":"'58ac023f','637870ab','12896cd4','689581dc','93f03899','880ad349','V86176471'",
    "Postpaid Bottom Banner":"'dd1a047d','a51dd0fd','8db773eb','48c34cca','1f1e35cb','572fab02'",
    "Prepaid Detail Offer Card":"'b9062d0d','a32fa3ce','b1af9808','c6a8a89e','237d525a','1e807957','1ddacf29','3d2b9c9b'",
    "BB Service Top Banner":"'f11054e7','2e186328','9fe9d099','d8e542ba','81abbd87','fcf1a079'",
    "BB Bottom Banner (dsl_service_banner)":"'6f74c144','a1299b91','f554e820','8091bc3d','2f4b8f6a','18256d1e'",
    "DTH Top Banner":"'b4c4e0c2','c4aa5d68','56f459c7','6a51a397','61af4976','9f072dbb'",
    "DTH Bottom Banner":"'991b44a7','409cd110','889e74d5','4d59d888','6819055d','320ef467'",
    "Bills Landing Offer Card":"'7306d6d3','a66f217b','d16811ed','63a763e2','1f32b96e','fc3d8172','0e71bb08','4131d235'",
    "Bills Offer Banner":"'b41a30df','c1c046f3','23574e26','019e0561','9e8de5b2','13dd0e50'",
    "Manage Offer Card":"'a6af3de2','0b5748d2','2225ac6c','7ef71759','a9768c28','246e961c','4b407082','3c474014','acf85d85','2b87f967','dbff6d13','5c80c9f1'",
    "Common Recharge Home":"'a1576aa8','9bc12867','f677b475','442af065','0f4baa33','7092b488','a3809593','4dc2e174'",
    "Prepaid Offer Recharge Screen":"'8b080012','02e62839','eb23acee','9dfb0a25','719e6100','f44e1be7','b7ba2f8b','a6fc9aa6'",
    "DTH Offer Recharge Screen":"'a7189ba7','6cc34096','2e7071d6','135c02d4','a636e489','7c1f771c','297a33de','a12d7202'",
    "Featuring Fresh":"'e06e3201','405af1e9','003bc238','7d101b9b','a94955f7','82a0e747','b9b44197','1486bc6f','496f352d','40d888e2','75e630d2','805411a4'",
    "Curated for you":"'f8e35b72','db1f535d','e204699a','ceda4da3','a298c073','869181a8','a6ce9d97','3ea51aec','ffb0d879','246911e7','52410df0','314d1738'",
    "FS_Explore_Money":"'66fd8c8b','68d64fc8','07260a1d','be7ac48a','2e444c8f','c868e58b'",
    "FS Pay Page":"'7c9471df','e7b53a02'",
    "FS Loan Page":"'f83d93bd','9fd39b6d'",
    "P2P Scratch Card":"'202ae67c','ff919627'"
}
# Current date and time
today = datetime.now()

# 2 weeks ago
two_weeks_ago = today - timedelta(weeks=7)
output = []

for inventory_name, inventory_adspot in inventory_metadata.items():
    # SQL query
    #query = """
    #SELECT FLOOR(__time TO DAY) as data_day,campaign_id, sum("requests") as request, sum("impressions") as impression, sum("clicks") as click
    #FROM a207953
    #WHERE __time >= '"""+two_weeks_ago.strftime("%Y-%m-%d")+"""' AND __time < '"""+today.strftime("%Y-%m-%d")+"""'
    #  AND adspot_id IN ("""+inventory_adspot+""")
    #GROUP BY FLOOR(__time TO DAY),campaign_id
    #"""

    query = """
    SELECT FLOOR(__time TO DAY) as data_day, sum("requests") as request, sum("ads") as ad, sum("impressions") as impression, sum("clicks") as click
    FROM a207953
    WHERE __time >= '"""+two_weeks_ago.strftime("%Y-%m-%d")+"""' AND __time < '"""+today.strftime("%Y-%m-%d")+"""'
      AND adspot_id IN ("""+inventory_adspot+""")
    GROUP BY FLOOR(__time TO DAY)
    """

    print(query)
    #exit(0)
    # Payload
    payload = {
        "query": query
    }

    # Send POST request
    response = requests.post(
        url,
        json=payload,
        auth=HTTPBasicAuth(username, password),
        headers={"Content-Type": "application/json"}
    )

    # Print results
    if response.status_code == 200:
        print("✅ Query executed successfully!")
        druid_response = response.json()
        for record in druid_response:
            record["inventory"] = inventory_name
            record["fillrate"] = (record["ad"]/record["request"])*100
            record["data_day"] = datetime.fromisoformat(record["data_day"].replace("Z", "+00:00"))
        #print(druid_response)
        #exit(0)
        output = output + druid_response
    else:
        print(f"❌ Error {response.status_code}: {response.text}")

#print(output)
# Write to CSV
#with open('inventory_report/inventory-wise-'+today.strftime("%Y-%m-%d")+'.csv', 'w', newline='') as csvfile:
#    fieldnames = output[0].keys()
#    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

#    writer.writeheader()  # Write CSV header
#    writer.writerows(output)  # Write all rows

# Sort data by inventory, then by date
output.sort(key=lambda x: (x["inventory"], x["data_day"]))

# Create Excel workbook
wb = Workbook()
ws = wb.active
ws.title = "Metrics"

# Define header
headers = ["data_day", "inventory", "request", "ad", "fillrate", "impression", "click"]
ws.append(headers)

# Define fills
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Increase
red_fill   = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")  # Decrease

# Track previous values per inventory
prev_values = {}

# Write rows with conditional formatting
for row in output:
    excel_row = [
        row["data_day"].strftime("%Y-%m-%d"),
        row["inventory"],
        row["request"],
        row["ad"],
        row["fillrate"],
        row["impression"],
        row["click"]
    ]
    ws.append(excel_row)

    # Current row index in Excel
    row_idx = ws.max_row

    # Check if we have previous values for this inventory
    inv = row["inventory"]
    if inv in prev_values:
        for col, metric in enumerate(["request", "ad", "fillrate", "impression", "click"], start=3):  # Excel col 3,4,5
            prev_val = prev_values[inv][metric]
            curr_val = row[metric]

            if curr_val > prev_val:
                ws.cell(row=row_idx, column=col).fill = green_fill
            elif curr_val < prev_val:
                ws.cell(row=row_idx, column=col).fill = red_fill
            # if equal → leave default

    # Update last seen values
    prev_values[inv] = row

# Save Excel
wb.save('inventory_report/inventory-wise-'+today.strftime("%Y-%m-%d")+'.xlsx')

print("✅ inventory_report/inventory-wise-"+today.strftime("%Y-%m-%d")+".xlsx file created successfully!")


subject = "Inventory wise report for last 7 weeks"
body = "Hi,\n\nPlease find the attached file.\n\nThanks"
msg = EmailMessage()
msg["From"] = EMAIL_SENDER
msg["To"] = ", ".join(EMAIL_RECIPIENTS) if isinstance(EMAIL_RECIPIENTS, (list, tuple)) else EMAIL_RECIPIENTS
msg["Subject"] = subject
msg.set_content(body)

# Attach file (binary-safe)
p = Path('inventory_report/inventory-wise-'+today.strftime("%Y-%m-%d")+'.xlsx')
with p.open("rb") as f:
    file_data = f.read()
msg.add_attachment(file_data, maintype="application", subtype="octet-stream", filename=p.name)

# Mail Send via SMTP
try:
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as s:
        s.ehlo()  # Say hello to the server
        s.send_message(msg)
    print("✅ Report email sent successfully!")
except Exception as e:
    print("❌ Error sending email:", e)

print(datetime.now().strftime('%Y-%m-%d %H:%M:%S'),"-------------Completed-------------")
